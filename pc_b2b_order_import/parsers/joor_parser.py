# -*- coding: utf-8 -*-
"""
Parser para ficheros Excel exportados desde la plataforma JOOR.

Formato JOOR:
- Una hoja con nombre "PO# XXXXXXXX"
- Cabecera (filas 2-16): marca, representante, PO number, cliente, dirección,
  fechas envío, condiciones pago
- Fila 18: cabeceras de las columnas de producto
  - Columnas fijas: A=Style Image, B=Style Name, C=Style Number, D=Color,
    E=Color Code, F=Color Comment, G=Style Comment, H=Materials, I=Fabrication,
    J=Country of Origin
  - Columnas dinámicas de tallas (K en adelante): nombres de tallas como cabecera
  - Últimas columnas: Sugg. Retail (EUR), WholeSale (EUR), Item Discount,
    Units, Total (EUR)
- A partir de fila 19: líneas de producto
- Tras las líneas de producto: fila de totales (Total:) y resumen
  (Subtotal, Discount, Shipping, Tax, Total)

Particularidades:
- Las tallas varían entre pedidos (35-42 para zapatos, 34-48+S/M/L/XL para ropa)
- La marca se extrae de celda B2
- El PO number de celda F4
- Las fechas de envío de celdas B15 y B16 (formato texto)
- Las cantidades están en las columnas de tallas (cero o vacío = no se pide)
"""
import logging
import re

from io import BytesIO

import openpyxl

_logger = logging.getLogger(__name__)


def parse_joor_file(file_content):
    """
    Parsea un fichero JOOR (.xlsx) y retorna un diccionario con la información
    del pedido y las líneas de producto.

    :param file_content: bytes del fichero Excel
    :return: dict con claves:
        - 'po_number': str
        - 'brand': str (marca/proveedor)
        - 'status': str
        - 'created_date': str
        - 'customer_id': str
        - 'customer_name': str
        - 'ship_to_address': dict
        - 'billing_address': dict
        - 'start_ship': str
        - 'complete_ship': str
        - 'payment_terms': str
        - 'carrier': str
        - 'lines': list of dict (líneas de producto)
        - 'totals': dict (subtotal, discount, shipping, tax, total)
    """
    wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)

    # Localizar la hoja PO# — puede ser la primera hoja o buscar por nombre
    po_sheet = None
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith('PO#') or sheet_name.startswith('PO #'):
            po_sheet = wb[sheet_name]
            break
    if po_sheet is None:
        # Usar la primera hoja si no encontramos una con PO#
        po_sheet = wb[wb.sheetnames[0]]

    ws = po_sheet
    result = {
        'po_number': '',
        'brand': '',
        'status': '',
        'created_date': '',
        'customer_id': '',
        'customer_name': '',
        'ship_to_address': {},
        'billing_address': {},
        'start_ship': '',
        'complete_ship': '',
        'payment_terms': '',
        'carrier': '',
        'lines': [],
        'totals': {},
    }

    # -------------------------------------------------------------------------
    # Extraer cabecera
    # -------------------------------------------------------------------------
    # B2 = Marca/Brand
    result['brand'] = _safe_str(ws, 'B2')

    # E2/F2 = Order Status
    result['status'] = _safe_str(ws, 'F2')

    # F3 = Created Date
    result['created_date'] = _safe_str(ws, 'F3')

    # F4 = PO Number
    result['po_number'] = _safe_str(ws, 'F4')

    # F5 = Customer ID
    result['customer_id'] = _safe_str(ws, 'F5')

    # B8 = Customer name
    result['customer_name'] = _safe_str(ws, 'B8')

    # Shipping address (D8-D12)
    result['ship_to_address'] = {
        'name': _safe_str(ws, 'D9'),
        'street': _safe_str(ws, 'D10'),
        'city_zip': _safe_str(ws, 'D11'),
        'country': _safe_str(ws, 'D12'),
    }

    # Billing address (F8-F12)
    result['billing_address'] = {
        'name': _safe_str(ws, 'F8'),
        'street': _safe_str(ws, 'F9') or _safe_str(ws, 'F10'),
        'city': _safe_str(ws, 'F10') or _safe_str(ws, 'F11'),
        'country': _safe_str(ws, 'F11') or _safe_str(ws, 'F12'),
    }

    # Dates (B15, B16) — formato texto "Start Ship: MM/DD/YYYY"
    start_ship_raw = _safe_str(ws, 'B15')
    if 'Start Ship:' in start_ship_raw:
        result['start_ship'] = start_ship_raw.replace('Start Ship:', '').strip()
    else:
        result['start_ship'] = start_ship_raw

    complete_raw = _safe_str(ws, 'B16')
    if 'Complete:' in complete_raw:
        result['complete_ship'] = complete_raw.replace('Complete:', '').strip()
    else:
        result['complete_ship'] = complete_raw

    # Carrier (D15-D16)
    result['carrier'] = _safe_str(ws, 'D16') or _safe_str(ws, 'D15')

    # Payment terms (F15-F16)
    terms_code = _safe_str(ws, 'F15')
    terms_desc = _safe_str(ws, 'F16')
    result['payment_terms'] = terms_desc or terms_code

    # -------------------------------------------------------------------------
    # Localizar fila de cabeceras de producto
    # -------------------------------------------------------------------------
    header_row = _find_header_row(ws)
    if header_row is None:
        raise ValueError(
            "No se ha encontrado la fila de cabeceras de producto en el fichero JOOR. "
            "Se buscaba 'Style Number' en las primeras 30 filas."
        )

    # Leer cabeceras
    headers = {}
    size_columns = []  # Lista de (col_index, size_name)

    for cell in ws[header_row]:
        if cell.value is None:
            continue
        col_letter = cell.column_letter
        col_idx = cell.column
        val = str(cell.value).strip()

        if val == 'Style Name':
            headers['style_name_col'] = col_idx
        elif val == 'Style Number':
            headers['style_number_col'] = col_idx
        elif val == 'Color':
            headers['color_col'] = col_idx
        elif val == 'Color Code':
            headers['color_code_col'] = col_idx
        elif val == 'Color Comment':
            headers['color_comment_col'] = col_idx
        elif val == 'Style Comment':
            headers['style_comment_col'] = col_idx
        elif val == 'Materials':
            headers['materials_col'] = col_idx
        elif val == 'Fabrication':
            headers['fabrication_col'] = col_idx
        elif val == 'Country of Origin':
            headers['country_col'] = col_idx
        elif val.startswith('Sugg. Retail'):
            headers['retail_price_col'] = col_idx
        elif val.startswith('WholeSale'):
            headers['wholesale_price_col'] = col_idx
        elif val == 'Item Discount':
            headers['discount_col'] = col_idx
        elif val == 'Units':
            headers['units_col'] = col_idx
        elif val.startswith('Total'):
            headers['total_col'] = col_idx
        elif val == 'Style Image':
            headers['image_col'] = col_idx
        else:
            # Si no es una cabecera conocida, es una talla
            # Las tallas están entre Country of Origin y Sugg. Retail
            size_columns.append((col_idx, val))

    # Filtrar size_columns: solo las que están entre country_col y retail_price_col
    country_col = headers.get('country_col', 0)
    retail_col = headers.get('retail_price_col', 999)
    size_columns = [
        (idx, name) for idx, name in size_columns
        if country_col < idx < retail_col
    ]

    # Limpiar prefijo underscore de tallas ASPESI: "__38" -> "38", "___L" -> "L"
    size_columns = [
        (idx, name.lstrip('_')) for idx, name in size_columns
    ]

    _logger.info(
        "JOOR Parser: header_row=%s, tallas=%s",
        header_row,
        [s[1] for s in size_columns],
    )

    # -------------------------------------------------------------------------
    # Leer líneas de producto
    # -------------------------------------------------------------------------
    line_num = 0
    for row_idx in range(header_row + 1, ws.max_row + 1):
        # Detectar fin de líneas: fila de totales ("Total:" en columna J)
        cell_j = ws.cell(row=row_idx, column=headers.get('country_col', 10))
        if cell_j.value and str(cell_j.value).strip() == 'Total:':
            break

        # Leer style_number — si está vacío, puede ser una fila de separación
        style_number_col = headers.get('style_number_col', 3)
        style_number = ws.cell(row=row_idx, column=style_number_col).value
        if not style_number:
            # Verificar si es una fila de grupo (ej. "JOFRE INTERNATIONAL - 000")
            # Estas filas solo tienen valor en columna A
            continue

        style_name = ws.cell(
            row=row_idx,
            column=headers.get('style_name_col', 2),
        ).value or ''
        color = ws.cell(
            row=row_idx,
            column=headers.get('color_col', 4),
        ).value or ''
        color_code = ws.cell(
            row=row_idx,
            column=headers.get('color_code_col', 5),
        ).value or ''
        materials = ws.cell(
            row=row_idx,
            column=headers.get('materials_col', 8),
        ).value or ''
        fabrication = ws.cell(
            row=row_idx,
            column=headers.get('fabrication_col', 9),
        ).value or ''
        country = ws.cell(
            row=row_idx,
            column=headers.get('country_col', 10),
        ).value or ''

        retail_price = _safe_float(
            ws.cell(row=row_idx, column=headers.get('retail_price_col', 0)).value
        )
        wholesale_price = _safe_float(
            ws.cell(row=row_idx, column=headers.get('wholesale_price_col', 0)).value
        )
        discount = _safe_float(
            ws.cell(row=row_idx, column=headers.get('discount_col', 0)).value
        )
        total_units = _safe_float(
            ws.cell(row=row_idx, column=headers.get('units_col', 0)).value
        )

        # Leer cantidades por talla
        for size_col_idx, size_name in size_columns:
            qty = _safe_float(ws.cell(row=row_idx, column=size_col_idx).value)
            if qty > 0:
                line_num += 1
                result['lines'].append({
                    'line_number': line_num,
                    'style_name': str(style_name).strip(),
                    'style_number': str(style_number).strip(),
                    'color': str(color).strip(),
                    'color_code': str(color_code).strip(),
                    'size': str(size_name).strip(),
                    'quantity': qty,
                    'wholesale_price': wholesale_price,
                    'retail_price': retail_price,
                    'discount': discount,
                    'materials': str(materials).strip(),
                    'fabrication': str(fabrication).strip(),
                    'country_of_origin': str(country).strip(),
                })

    # -------------------------------------------------------------------------
    # Leer totales (filas después de la fila Total:)
    # -------------------------------------------------------------------------
    # Buscar en las últimas filas
    for row_idx in range(ws.max_row, max(ws.max_row - 10, 1), -1):
        for cell in ws[row_idx]:
            if cell.value and 'Subtotal:' in str(cell.value):
                # El valor está en la columna siguiente o la última
                next_col = cell.column + 1
                result['totals']['subtotal'] = _safe_float(
                    ws.cell(row=row_idx, column=next_col).value
                )
            elif cell.value and str(cell.value).strip() == 'Total:':
                next_col = cell.column + 1
                val = ws.cell(row=row_idx, column=next_col).value
                if val and isinstance(val, (int, float)):
                    result['totals']['total'] = float(val)

    _logger.info(
        "JOOR Parser: Marca=%s, PO=%s, líneas=%d",
        result['brand'],
        result['po_number'],
        len(result['lines']),
    )

    return result


def _find_header_row(ws):
    """Busca la fila que contiene 'Style Number' como cabecera."""
    for row_idx in range(1, min(ws.max_row + 1, 30)):
        for cell in ws[row_idx]:
            if cell.value and str(cell.value).strip() == 'Style Number':
                return row_idx
    return None


def _safe_str(ws, cell_ref):
    """Lee una celda como string de forma segura."""
    val = ws[cell_ref].value
    if val is None:
        return ''
    return str(val).strip()


def _safe_float(value):
    """Convierte un valor a float de forma segura."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        # Intentar parsear strings con comas europeas
        cleaned = str(value).replace(',', '.').replace(' ', '')
        return float(cleaned)
    except (ValueError, TypeError):
        return 0.0
