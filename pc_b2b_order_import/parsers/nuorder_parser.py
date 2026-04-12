# -*- coding: utf-8 -*-
"""
Parser para ficheros Excel exportados desde la plataforma NuORDER.

Formato NuORDER:
- Hoja "NuORDER Order Data" con líneas de producto
  - Fila 1: cabeceras
  - Columnas fijas (A-S):
    A=Season, B=Image, C=Style Number, D=Color Code, E=Description,
    F=Name, G=Fabric Description, H=Color, I=Wholesale (EUR),
    J=M.S.R.P. (EUR), K=Division, L=Department, M=Category,
    N=Subcategory, O=Product Notes, P=Ship Start, Q=Ship End,
    R=Total Price (EUR), S=Total Units
  - Columnas dinámicas de tallas (T en adelante): patrón repetitivo
    Size N, Qty N, Size price N (de 1 a 17 o más)
  - Cada grupo de 3 columnas = (nombre talla, cantidad, precio unitario)

- Hoja "Summary" (solo informativa, no se importa):
  - Fila 1: "THIS TAB IS INFORMATIONAL ONLY"
  - Contiene Total Units y Order Total

Particularidades:
- Las tallas pueden ser numéricas (34.5, 35, 36...) o texto (PZ, S, M, L...)
- Una línea puede tener qty vacío o '' para una talla = no se pide
- Ship Start/Ship End son datetime
- Season es texto como "26W" (2026 Winter)
"""
import logging
from datetime import datetime
from io import BytesIO

import openpyxl

_logger = logging.getLogger(__name__)


def parse_nuorder_file(file_content):
    """
    Parsea un fichero NuORDER (.xlsx) y retorna un diccionario con las líneas
    de producto y metadatos.

    :param file_content: bytes del fichero Excel
    :return: dict con claves:
        - 'season': str
        - 'lines': list of dict (líneas individuales talla x producto)
        - 'ship_start': str
        - 'ship_end': str
        - 'summary': dict (total_units, order_total)
    """
    wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)

    # Localizar hoja de datos
    data_sheet = None
    for sheet_name in wb.sheetnames:
        if 'order data' in sheet_name.lower() or 'nuorder' in sheet_name.lower():
            data_sheet = wb[sheet_name]
            break
    if data_sheet is None:
        data_sheet = wb[wb.sheetnames[0]]

    ws = data_sheet
    result = {
        'season': '',
        'lines': [],
        'ship_start': '',
        'ship_end': '',
        'summary': {},
    }

    # -------------------------------------------------------------------------
    # Leer cabeceras (fila 1)
    # -------------------------------------------------------------------------
    headers = {}
    size_groups = []  # Lista de (size_col, qty_col, price_col, group_number)

    for cell in ws[1]:
        if cell.value is None:
            continue
        col_idx = cell.column
        val = str(cell.value).strip()

        # Cabeceras fijas
        if val == 'Season':
            headers['season_col'] = col_idx
        elif val == 'Style Number':
            headers['style_number_col'] = col_idx
        elif val == 'Color Code':
            headers['color_code_col'] = col_idx
        elif val == 'Description':
            headers['description_col'] = col_idx
        elif val == 'Name':
            headers['name_col'] = col_idx
        elif val == 'Fabric Description':
            headers['fabric_col'] = col_idx
        elif val == 'Color':
            headers['color_col'] = col_idx
        elif val.startswith('Wholesale'):
            headers['wholesale_col'] = col_idx
        elif val.startswith('M.S.R.P'):
            headers['retail_col'] = col_idx
        elif val == 'Division':
            headers['division_col'] = col_idx
        elif val == 'Department':
            headers['department_col'] = col_idx
        elif val == 'Category':
            headers['category_col'] = col_idx
        elif val == 'Subcategory':
            headers['subcategory_col'] = col_idx
        elif val == 'Product Notes':
            headers['notes_col'] = col_idx
        elif val == 'Ship Start':
            headers['ship_start_col'] = col_idx
        elif val == 'Ship End':
            headers['ship_end_col'] = col_idx
        elif val == 'Total Price (EUR)':
            headers['total_price_col'] = col_idx
        elif val == 'Total Units':
            headers['total_units_col'] = col_idx
        elif val.startswith('Size ') and not val.startswith('Size price'):
            # Extraer número de grupo
            match_num = val.replace('Size ', '')
            try:
                group_num = int(match_num)
                size_groups.append({
                    'size_col': col_idx,
                    'qty_col': col_idx + 1,  # Qty N está justo después de Size N
                    'price_col': col_idx + 2,  # Size price N después de Qty N
                    'group_num': group_num,
                })
            except ValueError:
                pass

    _logger.info(
        "NuORDER Parser: %d cabeceras fijas, %d grupos de tallas",
        len(headers),
        len(size_groups),
    )

    # -------------------------------------------------------------------------
    # Leer líneas de producto
    # -------------------------------------------------------------------------
    line_num = 0
    for row_idx in range(2, ws.max_row + 1):
        style_number = _cell_val(ws, row_idx, headers.get('style_number_col', 3))
        if not style_number:
            continue

        season = _cell_val(ws, row_idx, headers.get('season_col', 1)) or ''
        color_code = _cell_val(ws, row_idx, headers.get('color_code_col', 4)) or ''
        description = _cell_val(ws, row_idx, headers.get('description_col', 5)) or ''
        name = _cell_val(ws, row_idx, headers.get('name_col', 6)) or ''
        fabric = _cell_val(ws, row_idx, headers.get('fabric_col', 7)) or ''
        color = _cell_val(ws, row_idx, headers.get('color_col', 8)) or ''
        wholesale = _safe_float(
            _cell_val(ws, row_idx, headers.get('wholesale_col', 9))
        )
        retail = _safe_float(
            _cell_val(ws, row_idx, headers.get('retail_col', 10))
        )
        division = _cell_val(ws, row_idx, headers.get('division_col', 11)) or ''
        department = _cell_val(ws, row_idx, headers.get('department_col', 12)) or ''
        category = _cell_val(ws, row_idx, headers.get('category_col', 13)) or ''
        subcategory = _cell_val(ws, row_idx, headers.get('subcategory_col', 14)) or ''
        notes = _cell_val(ws, row_idx, headers.get('notes_col', 15)) or ''

        ship_start = _cell_val(ws, row_idx, headers.get('ship_start_col', 16))
        ship_end = _cell_val(ws, row_idx, headers.get('ship_end_col', 17))

        # Guardar season y fechas globales del primer registro
        if not result['season'] and season:
            result['season'] = str(season).strip()
        if not result['ship_start'] and ship_start:
            result['ship_start'] = _format_date(ship_start)
        if not result['ship_end'] and ship_end:
            result['ship_end'] = _format_date(ship_end)

        # Leer tallas y cantidades
        for sg in size_groups:
            size_name = _cell_val(ws, row_idx, sg['size_col'])
            qty_raw = _cell_val(ws, row_idx, sg['qty_col'])
            qty = _safe_float(qty_raw)

            if not size_name or qty <= 0:
                continue

            size_name = str(size_name).strip()
            line_num += 1

            # Precio por talla (puede diferir del wholesale global)
            size_price = _safe_float(
                _cell_val(ws, row_idx, sg['price_col'])
            )
            effective_price = size_price if size_price > 0 else wholesale

            result['lines'].append({
                'line_number': line_num,
                'season': str(season).strip(),
                'style_number': str(style_number).strip(),
                'style_name': str(name).strip() or str(description).strip(),
                'description': str(description).strip(),
                'color': str(color).strip(),
                'color_code': str(color_code).strip(),
                'size': size_name,
                'quantity': qty,
                'wholesale_price': effective_price,
                'retail_price': retail,
                'fabric': str(fabric).strip(),
                'division': str(division).strip(),
                'department': str(department).strip(),
                'category': str(category).strip(),
                'subcategory': str(subcategory).strip(),
                'notes': str(notes).strip(),
            })

    # -------------------------------------------------------------------------
    # Leer Summary (si existe)
    # -------------------------------------------------------------------------
    summary_sheet = None
    for sheet_name in wb.sheetnames:
        if 'summary' in sheet_name.lower():
            summary_sheet = wb[sheet_name]
            break

    if summary_sheet:
        for row_idx in range(1, summary_sheet.max_row + 1):
            cell_a = summary_sheet.cell(row=row_idx, column=1).value
            cell_b = summary_sheet.cell(row=row_idx, column=2).value
            if cell_a and 'Total Units' in str(cell_a):
                result['summary']['total_units'] = _safe_float(cell_b)
            elif cell_a and 'Order Total' in str(cell_a):
                result['summary']['order_total'] = _safe_float(cell_b)

    _logger.info(
        "NuORDER Parser: season=%s, líneas=%d",
        result['season'],
        len(result['lines']),
    )

    return result


def _cell_val(ws, row, col):
    """Lee el valor de una celda de forma segura."""
    if col is None or col <= 0:
        return None
    val = ws.cell(row=row, column=col).value
    if val == '' or val is None:
        return None
    return val


def _safe_float(value):
    """Convierte un valor a float de forma segura."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        cleaned = str(value).replace(',', '.').replace(' ', '')
        return float(cleaned)
    except (ValueError, TypeError):
        return 0.0


def _format_date(value):
    """Formatea una fecha como string DD/MM/YYYY."""
    if isinstance(value, datetime):
        return value.strftime('%d/%m/%Y')
    return str(value).strip() if value else ''
